import openpyxl
from DatadrivenTesting import XLUtils


import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
options = webdriver.ChromeOptions()
options.add_argument("--disable-notifications")  # By using this we can ignore browser notification popUps


options.add_experimental_option("detach", True)
srv_obj = Service(r"C:\pythonProject\venv\Scripts\chromedriver")
driver = webdriver.Chrome(service=srv_obj, options=options)
driver.implicitly_wait(10)


file = "C:\\Users\\Administrator\\Documents\\samplexl.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]  # Get the active sheet in Excel
sheet = workbook["Sheet1"]
rows = sheet.max_row  # count number of rows in a Excel sheet
cols = sheet.max_column  # count number of columns in a Excel sheet



for r in range(2, rows+1):
    pric = XLUtils.readData(file,"Sheet1",r,1)
    RatioOfInterest = XLUtils.readData(file,"Sheet1",r,2)
    per1 = XLUtils.readData(file,"Sheet1",r,3)
    per2 = XLUtils.readData(file,"Sheet1",r,4)
    fre = XLUtils.readData(file,"Sheet1",r,5)
    exp_mvalue = XLUtils.readData(file,"Sheet1",r,6)

    driver.get("https://www.moneycontrol.com/fixed-income/calculator/state-bank-of-india-sbi/fixed-deposit-calculator-SBI-BSB001.html?classic=true")
    driver.maximize_window()

    driver.find_element(By.XPATH,"//input[@id='principal']").send_keys(pric)
    driver.find_element(By.XPATH,"//input[@id='interest']").send_keys(RatioOfInterest)
    driver.find_element(By.XPATH, "//input[@id='tenure']").send_keys(per1)
    priod_drop =Select(driver.find_element(By.XPATH,"//select[@id='tenurePeriod']"))
    priod_drop.select_by_visible_text(per2)
    frq_drop =Select(driver.find_element(By.XPATH,"//select[@id='frequency']"))
    frq_drop.select_by_visible_text(fre)
    # cal_button
    driver.find_element(By.XPATH, "/html/body/div[6]/div[2]/div/div[5]/div/div[1]/div[3]/form/div[2]/a[1]/img").click()
    actul_mvalue = driver.find_element(By.XPATH,"//span/strong").text

    #validation
    if float(exp_mvalue)==float(actul_mvalue):
        print("test passed")
        XLUtils.writeData(file,"Sheet1", r, 8, "Passed")
        XLUtils.fillGreenColor(file,"Sheet1", r,8)
    else:
        print("test failed")
        XLUtils.writeData(file, "Sheet1", r, 8, "Fail")
        XLUtils.fillRedColor(file, "Sheet1", r, 8)
    #clear button click
    driver.find_element(By.XPATH,"//body/div[@id='mc_mainWrapper']/div[2]/div[1]/div[5]/div[1]/div[1]/div[3]/form[1]/div[2]/a[2]/img[1]").click()
    time.sleep(5)
driver.quit()