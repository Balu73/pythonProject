import openpyxl

# Read the data in Excel file we need to follow the steps:

# File--> Workbook--> Sheet--> Rows--> Cells
file = r"C:\Users\Administrator\Documents\samplexl.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]
rows = sheet.max_row  # count number of rows in a Excel sheet
cols = sheet.max_column  # count number of columns in a Excel sheet

# Reading all the rows & columns from Excel sheet
for r in range(1, rows + 1):
    for c in range(1, cols + 1):
        print(sheet.cell(r, c).value, end="                       ")
    print()