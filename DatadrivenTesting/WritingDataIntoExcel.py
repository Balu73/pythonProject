import openpyxl

file = "C:\\Users\\Administrator\\Documents\\writing data.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]  # Get the active sheet in Excel

# If you want same data in all give cells then use the following code

# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value= "wellcome"
# workbook.save(file)  # this method used to we can save the data in the Excel sheet


# If you want give the data dynamically in the cells then use the following code
for r in range(1,4):
    for c in range(1,4):
        sheet.cell(r,c).value= input(f"please enter the data in cell({r, c}) :")
        print(sheet.cell(r,c).value, end="         ")
    print()
workbook.save(file)


# multiple data

# first row data
# sheet.cell(1, 1).value = 1001
# sheet.cell(1, 2).value = "smith"
# sheet.cell(1, 3).value = "engineer"
#
# # second row data
#
# sheet.cell(2, 1).value = 1002
# sheet.cell(2, 2).value = "william"
# sheet.cell(2, 3).value = "tester"
#
# # third row data
#
# sheet.cell(3, 1).value = 1003
# sheet.cell(3, 2).value = "john"
# sheet.cell(3, 3).value = "developer"
# workbook.save(file)